#!/usr/bin/env python3
"""
Daysee grant scraper with title-driven official URL resolution.

Design goals
------------
1. Stable crawl of the Daysee listing and detail pages.
2. Treat `title` as the primary key for URL resolution.
3. Resolve organizer/official URLs using:
   - verified direct links from the page when available
   - government-domain-constrained search (Bing HTML, optional DDG fallback)
   - local cache of previously verified title -> URL mappings
4. Never invent government domains by string concatenation.
5. Verify candidate URLs before writing them into Excel.
6. Produce both a full workbook and a delta workbook.

Environment variables
---------------------
OUTPUT_XLSX   default: outputs/daysee_grants.xlsx
DELTA_XLSX    default: outputs/daysee_grants_delta_only.xlsx
PREVIOUS_XLSX default: state/daysee_grants_latest.xlsx
CACHE_JSON    default: state/title_url_cache.json
MAX_PAGES     default: 25
REQUEST_TIMEOUT default: 20

Notes
-----
- This script avoids paid APIs.
- Search is best-effort. For future new plans, verification + cache make results
  more stable than scraping raw Google search URLs, but there can still be edge cases.
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import os
import re
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qs, quote_plus, unquote_plus, urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from pydantic import BaseModel, Field, HttpUrl, field_validator

# -------------------------
# Configuration
# -------------------------
BASE_URL = "https://dayseechat.com"
LISTING_URL = f"{BASE_URL}/explore-grant/"
OUTPUT_XLSX = os.getenv("OUTPUT_XLSX", "outputs/daysee_grants.xlsx")
DELTA_XLSX = os.getenv("DELTA_XLSX", "outputs/daysee_grants_delta_only.xlsx")
PREVIOUS_XLSX = os.getenv("PREVIOUS_XLSX", "state/daysee_grants_latest.xlsx")
CACHE_JSON = os.getenv("CACHE_JSON", "state/title_url_cache.json")
MAX_PAGES = int(os.getenv("MAX_PAGES", "25"))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "20"))

Path(OUTPUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
Path(DELTA_XLSX).parent.mkdir(parents=True, exist_ok=True)
Path(PREVIOUS_XLSX).parent.mkdir(parents=True, exist_ok=True)
Path(CACHE_JSON).parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
SEARCH_EXCLUDE_DOMAINS = {
    "google.com",
    "www.google.com",
    "support.google.com",
    "bing.com",
    "www.bing.com",
    "duckduckgo.com",
    "html.duckduckgo.com",
    "facebook.com",
    "www.facebook.com",
    "instagram.com",
    "www.instagram.com",
    "youtube.com",
    "www.youtube.com",
    "104.com.tw",
    "www.104.com.tw",
    "linkedin.com",
    "www.linkedin.com",
    "x.com",
    "twitter.com",
}

CENTRAL_SOURCE_HINTS: Dict[str, List[str]] = {
    "勞動部": ["www.mol.gov.tw", "wlb.mol.gov.tw", "wda.gov.tw"],
    "教育部": ["www.edu.tw", "depart.moe.edu.tw", "www.moe.gov.tw"],
    "經濟部": ["www.moea.gov.tw", "gcis.nat.gov.tw", "www.sme.gov.tw"],
    "數位發展部": ["www.moda.gov.tw", "moda.gov.tw", "digiplus.adi.gov.tw"],
    "客家委員會": ["www.hakka.gov.tw"],
    "原住民族委員會": ["www.cip.gov.tw"],
    "海洋委員會": ["www.oac.gov.tw", "oac.gov.tw"],
    "國家發展委員會": ["www.ndc.gov.tw", "ndc.gov.tw"],
    "文化部": ["www.moc.gov.tw", "grants.moc.gov.tw", "mocfile.moc.gov.tw"],
    "衛生福利部": ["www.mohw.gov.tw", "www.mohw.gov.tw/mp-1.html"],
    "農業部": ["www.moa.gov.tw", "agronet.zero.moa.gov.tw"],
    "環境部": ["www.moenv.gov.tw", "www.epa.gov.tw"],
    "交通部": ["www.motc.gov.tw", "www.taiwan.net.tw"],
}

LOCAL_GOV_HINTS: Dict[str, List[str]] = {
    "臺北": ["www.gov.taipei"],
    "台北": ["www.gov.taipei"],
    "新北": ["www.ntpc.gov.tw", "www.economic.ntpc.gov.tw", "www.culture.ntpc.gov.tw", "www.sw.ntpc.gov.tw"],
    "桃園": ["www.tycg.gov.tw", "youth.tycg.gov.tw"],
    "臺中": ["www.taichung.gov.tw", "www.economic.taichung.gov.tw", "www.culture.taichung.gov.tw"],
    "台中": ["www.taichung.gov.tw", "www.economic.taichung.gov.tw", "www.culture.taichung.gov.tw"],
    "臺南": ["www.tainan.gov.tw", "economic.tainan.gov.tw", "culture.tainan.gov.tw"],
    "台南": ["www.tainan.gov.tw", "economic.tainan.gov.tw", "culture.tainan.gov.tw"],
    "高雄": ["www.kcg.gov.tw", "edbkcg.kcg.gov.tw", "khh.travel", "youth.kcg.gov.tw"],
    "基隆": ["www.klcg.gov.tw"],
    "新竹市": ["www.hccg.gov.tw", "youthhsinchu.hccg.gov.tw"],
    "新竹縣": ["www.hsinchu.gov.tw"],
    "苗栗": ["www.miaoli.gov.tw"],
    "彰化": ["www.chcg.gov.tw", "www.changhua.gov.tw"],
    "南投": ["www.nantou.gov.tw"],
    "雲林": ["www.yunlin.gov.tw"],
    "嘉義市": ["www.chiayi.gov.tw"],
    "嘉義縣": ["www.cyhg.gov.tw"],
    "屏東": ["www.pthg.gov.tw"],
    "宜蘭": ["www.e-land.gov.tw"],
    "花蓮": ["www.hl.gov.tw"],
    "臺東": ["www.taitung.gov.tw"],
    "台東": ["www.taitung.gov.tw"],
    "澎湖": ["www.penghu.gov.tw"],
    "金門": ["www.kinmen.gov.tw"],
    "連江": ["www.matsu.gov.tw"],
    "馬祖": ["www.matsu.gov.tw"],
}

# -------------------------
# Models
# -------------------------
class ListingItem(BaseModel):
    title: str
    detail_url: HttpUrl
    plan_source: str = ""
    eligible_targets: str = ""
    applicable_region: str = ""
    grant_amount: str = ""
    deadline_date: str = ""
    deadline_text: str = ""
    topic_1: str = ""
    topic_2: str = ""
    topic_3: str = ""
    topic_4: str = ""
    topic_5: str = ""


class DetailItem(ListingItem):
    organizer_site_url_raw: str = ""
    organizer_search_query: str = ""
    organizer_search_scope: str = ""
    organizer_site_url: str = ""
    organizer_site_domain: str = ""
    official_organizer_site_url: str = ""
    official_organizer_domain: str = ""
    official_url_status: str = ""
    official_url_confidence: str = ""
    plan_background: str = ""
    key_point_1: str = ""
    key_point_2: str = ""
    key_point_3: str = ""
    key_point_4: str = ""
    key_point_5: str = ""
    application_tips: str = ""
    organizer_site_note: str = ""
    raw_text: str = ""

    @field_validator("detail_url", mode="before")
    @classmethod
    def ensure_detail_url(cls, value: Any) -> Any:
        return str(value)


@dataclass
class CandidateURL:
    url: str
    source: str
    scope: str
    query: str
    score: float = 0.0


# -------------------------
# Utility helpers
# -------------------------
def clean_text(text: Optional[str]) -> str:
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    return text


def split_lines(text: str) -> List[str]:
    return [clean_text(x) for x in re.split(r"[\r\n]+", text) if clean_text(x)]


def get_domain(url: str) -> str:
    if not url:
        return ""
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ""


def is_google_search_url(url: str) -> bool:
    d = get_domain(url)
    return "google." in d and "/search" in url


def extract_google_query(url: str) -> str:
    try:
        qs = parse_qs(urlparse(url).query)
        return unquote_plus(qs.get("q", [""])[0]).strip()
    except Exception:
        return ""


def is_probably_government_domain(domain: str) -> bool:
    d = domain.lower().strip()
    if not d:
        return False
    gov_markers = [".gov.tw", ".gov", "gov.taipei", "nat.gov.tw", "kcg.gov.tw", "hccg.gov.tw"]
    return any(m in d for m in gov_markers)


def normalize_title(title: str) -> Dict[str, str]:
    full = clean_text(title)
    core = full
    # remove leading year markers
    core = re.sub(r"^(\d{3,4}|20\d{2})年度?", "", core)
    core = re.sub(r"^(\d{3,4}|20\d{2})年", "", core)
    core = re.sub(r"^[（(【\[]?\d{3,4}[）)】\]]", "", core)
    core = re.sub(r"\b(徵件公告|補助要點|申請須知|計畫書|作業要點)$", "", core)
    core = clean_text(core)
    keywords = re.sub(r"[｜|、，,。．()（）【】\[\]：:]+", " ", core)
    keywords = re.sub(r"\s+", " ", keywords).strip()
    return {"full": full, "core": core or full, "keywords": keywords or full}


def title_key(title: str, source: str, region: str) -> str:
    norm = normalize_title(title)
    raw = f"{norm['core']}||{clean_text(source)}||{clean_text(region)}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def keyword_overlap_score(title: str, text: str) -> float:
    title_kw = {x for x in normalize_title(title)["keywords"].split(" ") if len(x) >= 2}
    body_kw = set(re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,}", text))
    if not title_kw:
        return 0.0
    inter = len(title_kw & body_kw)
    return inter / max(len(title_kw), 1)


def safe_request(method: str, url: str, **kwargs) -> Optional[requests.Response]:
    headers = kwargs.pop("headers", {}) or {}
    headers.setdefault("User-Agent", UA)
    headers.setdefault("Accept-Language", "zh-TW,zh;q=0.9,en;q=0.8")
    try:
        resp = requests.request(method, url, timeout=REQUEST_TIMEOUT, headers=headers, allow_redirects=True, **kwargs)
        return resp
    except Exception:
        return None


# -------------------------
# Search cache
# -------------------------
def load_cache() -> Dict[str, Dict[str, Any]]:
    path = Path(CACHE_JSON)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_cache(cache: Dict[str, Dict[str, Any]]) -> None:
    Path(CACHE_JSON).write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


# -------------------------
# Daysee parsing
# -------------------------
def extract_after_label(lines: Sequence[str], start_label: str, stop_labels: Sequence[str]) -> List[str]:
    capture = False
    out: List[str] = []
    for line in lines:
        if line.startswith(start_label):
            capture = True
            tail = clean_text(line.replace(start_label, "", 1))
            if tail:
                out.append(tail)
            continue
        if capture and any(line.startswith(x) for x in stop_labels):
            break
        if capture and line:
            out.append(line)
    # de-dup, preserve order
    seen = set()
    uniq = []
    for x in out:
        if x not in seen:
            uniq.append(x)
            seen.add(x)
    return uniq


def parse_total_results(html: str) -> int:
    text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
    m = re.search(r"(\d+)\s*out\s*of\s*(\d+)\s*results", text, re.I)
    if m:
        return int(m.group(2))
    return 0


def parse_listing_page(html: str, seen_urls: set[str]) -> List[ListingItem]:
    soup = BeautifulSoup(html, "html.parser")
    results: List[ListingItem] = []
    for a in soup.select('a[href*="/subsidy/grant-"]'):
        href = urljoin(BASE_URL, a.get("href", "").strip())
        if not href or href in seen_urls:
            continue
        title = clean_text(a.get_text(" ", strip=True))
        if not title:
            continue
        block = a
        for _ in range(8):
            txt = block.get_text("\n", strip=True) if hasattr(block, "get_text") else ""
            if "截止日期" in txt and ("計畫來源" in txt or "補助對象" in txt):
                break
            if getattr(block, "parent", None) is None:
                break
            block = block.parent
        lines = split_lines(block.get_text("\n", strip=True))
        topics = extract_after_label(lines, "＃關注議題：", ["＃補助對象：", "＃計畫來源：", "＃補助金額：", "截止日期："])
        recipients = extract_after_label(lines, "＃補助對象：", ["＃計畫來源：", "＃補助金額：", "截止日期："])
        sources = extract_after_label(lines, "＃計畫來源：", ["＃補助金額：", "截止日期："])
        amounts = extract_after_label(lines, "＃補助金額：", ["截止日期："])
        deadline_date = ""
        for line in lines:
            if line.startswith("截止日期："):
                m = re.search(r"(\d{4}-\d{2}-\d{2})", line)
                if m:
                    deadline_date = m.group(1)
                break
        item = ListingItem(
            title=title,
            detail_url=href,
            plan_source="｜".join(sources),
            eligible_targets="｜".join(recipients),
            applicable_region="",
            grant_amount="｜".join([x for x in amounts if not x.isdigit()]),
            deadline_date=deadline_date,
            deadline_text="",
            topic_1=topics[0] if len(topics) > 0 else "",
            topic_2=topics[1] if len(topics) > 1 else "",
            topic_3=topics[2] if len(topics) > 2 else "",
            topic_4=topics[3] if len(topics) > 3 else "",
            topic_5=topics[4] if len(topics) > 4 else "",
        )
        results.append(item)
        seen_urls.add(href)
    return results


def find_google_search_link(soup: BeautifulSoup) -> str:
    for a in soup.select("a[href]"):
        href = a.get("href", "").strip()
        if is_google_search_url(href):
            return href
    return ""


def parse_detail_page(html: str, detail_url: str) -> DetailItem:
    soup = BeautifulSoup(html, "html.parser")
    page_text = soup.get_text("\n", strip=True)
    lines = split_lines(page_text)

    title = ""
    for sel in ["h1", "h2", "title"]:
        node = soup.select_one(sel)
        if node and clean_text(node.get_text(" ", strip=True)):
            title = clean_text(node.get_text(" ", strip=True))
            break

    if title.endswith("｜小社區大事件"):
        title = clean_text(title.replace("｜小社區大事件", ""))

    def after(label: str, stops: Sequence[str]) -> List[str]:
        return extract_after_label(lines, label, stops)

    plan_source = "｜".join(after("計畫來源：", ["補助對象：", "適用地區：", "補助金額：", "截止日期：", "關注議題："]))
    eligible_targets = "｜".join(after("補助對象：", ["適用地區：", "補助金額：", "截止日期：", "關注議題："]))
    applicable_region = "｜".join(after("適用地區：", ["補助金額：", "截止日期：", "關注議題："]))
    grant_amount = "｜".join(after("補助金額：", ["截止日期：", "關注議題：", "申請文件：", "計畫背景："]))

    deadline_date = ""
    deadline_text = ""
    for line in lines:
        if line.startswith("截止日期："):
            payload = clean_text(line.replace("截止日期：", "", 1))
            m = re.search(r"(\d{4}-\d{2}-\d{2})", payload)
            if m:
                deadline_date = m.group(1)
                rest = payload.replace(deadline_date, "").strip(" ｜|")
                deadline_text = rest
            else:
                deadline_text = payload
            break

    topics = after("關注議題：", ["申請文件：", "計畫背景：", "計畫重點：", "撰寫建議："])
    plan_background = "\n".join(after("計畫背景：", ["計畫重點：", "撰寫建議：", "申請文件："]))
    points = after("計畫重點：", ["撰寫建議：", "申請文件："])
    tips = "\n".join(after("撰寫建議：", ["申請文件："]))

    direct_links: List[str] = []
    for a in soup.select("a[href]"):
        href = urljoin(BASE_URL, a.get("href", "").strip())
        txt = clean_text(a.get_text(" ", strip=True))
        if not href:
            continue
        domain = get_domain(href)
        if any(x in domain for x in ["google.com", "bing.com", "duckduckgo.com"]):
            continue
        if txt and any(k in txt for k in ["申請文件", "主辦單位", "官方", "計畫網站", "公告", "簡章", "附件"]):
            direct_links.append(href)

    organizer_site_url_raw = find_google_search_link(soup)
    if not organizer_site_url_raw and direct_links:
        organizer_site_url_raw = direct_links[0]

    item = DetailItem(
        title=title,
        detail_url=detail_url,
        plan_source=plan_source,
        eligible_targets=eligible_targets,
        applicable_region=applicable_region,
        grant_amount=grant_amount,
        organizer_site_url_raw=organizer_site_url_raw,
        deadline_date=deadline_date,
        deadline_text=deadline_text,
        topic_1=topics[0] if len(topics) > 0 else "",
        topic_2=topics[1] if len(topics) > 1 else "",
        topic_3=topics[2] if len(topics) > 2 else "",
        topic_4=topics[3] if len(topics) > 3 else "",
        topic_5=topics[4] if len(topics) > 4 else "",
        plan_background=plan_background,
        key_point_1=points[0] if len(points) > 0 else "",
        key_point_2=points[1] if len(points) > 1 else "",
        key_point_3=points[2] if len(points) > 2 else "",
        key_point_4=points[3] if len(points) > 3 else "",
        key_point_5=points[4] if len(points) > 4 else "",
        application_tips=tips,
        raw_text="\n".join(lines),
    )
    # Prefer verified direct link from page if there is one.
    if direct_links:
        item.organizer_site_url = direct_links[0]
        item.organizer_site_domain = get_domain(direct_links[0])
        if is_probably_government_domain(item.organizer_site_domain):
            item.official_organizer_site_url = direct_links[0]
            item.official_organizer_domain = item.organizer_site_domain
            item.official_url_status = "direct_official"
            item.official_url_confidence = "high"
        else:
            item.official_url_status = "direct_non_google"
            item.official_url_confidence = "medium"
    return item


# -------------------------
# Search + verification layer
# -------------------------
def build_domain_hints(title: str, plan_source: str, region: str) -> List[str]:
    domains: List[str] = []
    for key, vals in CENTRAL_SOURCE_HINTS.items():
        if key and key in plan_source:
            domains.extend(vals)
    haystack = " | ".join([title, plan_source, region])
    for key, vals in LOCAL_GOV_HINTS.items():
        if key and key in haystack:
            domains.extend(vals)
    if "縣市政府" in plan_source and not domains:
        for key, vals in LOCAL_GOV_HINTS.items():
            if key in title:
                domains.extend(vals)
    # de-dup preserve order
    seen = set()
    out = []
    for d in domains:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out


def search_bing_html(query: str) -> List[str]:
    url = f"https://www.bing.com/search?q={quote_plus(query)}&setlang=zh-Hant"
    resp = safe_request("GET", url)
    if not resp or resp.status_code != 200:
        return []
    soup = BeautifulSoup(resp.text, "html.parser")
    urls: List[str] = []
    for a in soup.select("li.b_algo h2 a[href], a[href]"):
        href = a.get("href", "").strip()
        if not href.startswith("http"):
            continue
        d = get_domain(href)
        if d in SEARCH_EXCLUDE_DOMAINS:
            continue
        if href not in urls:
            urls.append(href)
    return urls[:10]


def search_ddg_html(query: str) -> List[str]:
    url = "https://html.duckduckgo.com/html/"
    resp = safe_request("POST", url, data={"q": query})
    if not resp or resp.status_code != 200:
        return []
    soup = BeautifulSoup(resp.text, "html.parser")
    urls: List[str] = []
    for a in soup.select("a.result__a[href], a[href]"):
        href = a.get("href", "").strip()
        if not href.startswith("http"):
            continue
        d = get_domain(href)
        if d in SEARCH_EXCLUDE_DOMAINS:
            continue
        if href not in urls:
            urls.append(href)
    return urls[:10]


def verify_candidate(url: str, title: str, allowed_domains: Sequence[str]) -> Tuple[bool, str, float]:
    if not url:
        return False, "", 0.0
    domain = get_domain(url)
    if domain in SEARCH_EXCLUDE_DOMAINS:
        return False, domain, 0.0

    # Try variants to fix missing www issues.
    variants = [url]
    parsed = urlparse(url)
    if parsed.scheme in {"http", "https"} and parsed.netloc:
        host = parsed.netloc
        if host.startswith("www."):
            variants.append(url.replace("//www.", "//", 1))
        else:
            variants.append(url.replace(f"//{host}", f"//www.{host}", 1))

    best_score = -1.0
    best_url = ""
    for variant in variants:
        resp = safe_request("GET", variant, stream=False)
        if not resp or resp.status_code >= 400:
            continue
        final_url = resp.url
        final_domain = get_domain(final_url)
        content = resp.text[:120000]
        page_title = ""
        m = re.search(r"<title[^>]*>(.*?)</title>", content, re.I | re.S)
        if m:
            page_title = clean_text(BeautifulSoup(m.group(1), "html.parser").get_text(" ", strip=True))
        text_probe = f"{page_title}\n{BeautifulSoup(content, 'html.parser').get_text(' ', strip=True)[:8000]}"
        overlap = keyword_overlap_score(title, text_probe)
        gov_bonus = 0.35 if is_probably_government_domain(final_domain) else 0.0
        allowed_bonus = 0.35 if any(final_domain.endswith(x) or final_domain == x for x in allowed_domains) else 0.0
        path_bonus = 0.1 if re.search(r"(grant|subsid|news|bulletin|download|gov|plan|activity|article|content)", final_url, re.I) else 0.0
        score = overlap + gov_bonus + allowed_bonus + path_bonus
        if score > best_score:
            best_score = score
            best_url = final_url
    return best_score >= 0.55, get_domain(best_url), best_score if best_score > 0 else 0.0


def resolve_by_title(item: DetailItem, cache: Dict[str, Dict[str, Any]]) -> DetailItem:
    # Keep trustworthy direct links.
    if item.official_url_status in {"direct_official", "direct_non_google"}:
        return item

    key = title_key(item.title, item.plan_source, item.applicable_region)
    cached = cache.get(key)
    if cached:
        item.organizer_search_query = cached.get("query", "")
        item.organizer_search_scope = cached.get("scope", "")
        item.organizer_site_url = cached.get("organizer_site_url", item.organizer_site_url)
        item.organizer_site_domain = cached.get("organizer_site_domain", item.organizer_site_domain)
        item.official_organizer_site_url = cached.get("official_organizer_site_url", item.official_organizer_site_url)
        item.official_organizer_domain = cached.get("official_organizer_domain", item.official_organizer_domain)
        item.official_url_status = "cache_reused_verified"
        item.official_url_confidence = cached.get("confidence", "high")
        return item

    norm = normalize_title(item.title)
    hints = build_domain_hints(item.title, item.plan_source, item.applicable_region)

    queries: List[Tuple[str, str]] = []  # (query, scope)
    if hints:
        for d in hints[:4]:
            queries.append((f'"{norm["full"]}" site:{d}', d))
            if norm["core"] != norm["full"]:
                queries.append((f'"{norm["core"]}" site:{d}', d))
            queries.append((f'"{norm["core"]}" 補助 site:{d}', d))
            queries.append((f'"{norm["core"]}" 公告 site:{d}', d))
    else:
        queries.append((norm["full"], ""))
        if norm["core"] != norm["full"]:
            queries.append((norm["core"], ""))

    candidates: List[CandidateURL] = []
    seen_urls = set()
    for query, scope in queries[:8]:
        item.organizer_search_query = query
        item.organizer_search_scope = scope
        urls = search_bing_html(query)
        if not urls:
            urls = search_ddg_html(query)
        for u in urls:
            if u in seen_urls:
                continue
            seen_urls.add(u)
            candidates.append(CandidateURL(url=u, source="search", scope=scope, query=query))
        if len(candidates) >= 12:
            break
        time.sleep(0.3)

    best_verified: Optional[Tuple[str, str, float, str, str]] = None  # url, domain, score, query, scope
    for cand in candidates:
        ok, domain, score = verify_candidate(cand.url, item.title, hints)
        if ok:
            if best_verified is None or score > best_verified[2]:
                best_verified = (cand.url, domain, score, cand.query, cand.scope)

    if best_verified:
        url, domain, score, query, scope = best_verified
        item.organizer_search_query = query
        item.organizer_search_scope = scope
        item.organizer_site_url = url
        item.organizer_site_domain = domain
        if is_probably_government_domain(domain):
            item.official_organizer_site_url = url
            item.official_organizer_domain = domain
            item.official_url_status = "search_verified_exact"
            item.official_url_confidence = "high" if score >= 0.9 else "medium"
        else:
            item.official_url_status = "search_verified_non_gov"
            item.official_url_confidence = "medium"

        cache[key] = {
            "query": query,
            "scope": scope,
            "organizer_site_url": item.organizer_site_url,
            "organizer_site_domain": item.organizer_site_domain,
            "official_organizer_site_url": item.official_organizer_site_url,
            "official_organizer_domain": item.official_organizer_domain,
            "confidence": item.official_url_confidence,
        }
        return item

    # Last resort: keep original search URL if present; do not invent portals.
    raw = item.organizer_site_url_raw
    if raw:
        item.organizer_site_url = raw
        item.organizer_site_domain = get_domain(raw)
    item.official_url_status = "search_no_match"
    item.official_url_confidence = "low"
    return item


def listing_fallback(item: ListingItem) -> DetailItem:
    return DetailItem(
        title=item.title,
        detail_url=str(item.detail_url),
        plan_source=item.plan_source,
        eligible_targets=item.eligible_targets,
        applicable_region=item.applicable_region,
        grant_amount=item.grant_amount,
        deadline_date=item.deadline_date,
        deadline_text=item.deadline_text,
        topic_1=item.topic_1,
        topic_2=item.topic_2,
        topic_3=item.topic_3,
        topic_4=item.topic_4,
        topic_5=item.topic_5,
        official_url_status="detail_timeout_fallback",
        official_url_confidence="low",
    )


async def extract_detail(page, item: ListingItem) -> DetailItem:
    detail_url = str(item.detail_url)
    logger.info("Fetching detail: %s", detail_url)
    last_exc = None
    for attempt in range(1, 4):
        try:
            await page.goto(detail_url, wait_until="domcontentloaded", timeout=90000)
            await page.wait_for_timeout(1500)
            html = await page.content()
            return parse_detail_page(html, detail_url)
        except PlaywrightTimeoutError as exc:
            last_exc = exc
            logger.warning("Detail timeout for %s (attempt %d/3)", detail_url, attempt)
            try:
                await page.goto("about:blank", wait_until="load", timeout=10000)
            except Exception:
                pass
            await page.wait_for_timeout(1200 * attempt)
        except Exception as exc:
            last_exc = exc
            logger.warning("Detail fetch failed for %s (attempt %d/3): %s", detail_url, attempt, exc)
            try:
                await page.goto("about:blank", wait_until="load", timeout=10000)
            except Exception:
                pass
            await page.wait_for_timeout(1200 * attempt)
    logger.error("Fallback to listing-only detail for %s because detail page could not be read: %s", detail_url, last_exc)
    return listing_fallback(item)


async def extract_all_listings(page) -> List[ListingItem]:
    await page.goto(LISTING_URL, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(2500)

    seen_urls: set[str] = set()
    all_items: List[ListingItem] = []
    target_total = 0
    stagnation = 0

    for page_no in range(1, MAX_PAGES + 1):
        logger.info("Scanning listing page %d", page_no)
        html = await page.content()
        if not target_total:
            target_total = parse_total_results(html)
            if target_total:
                logger.info("Target total grants detected: %d", target_total)

        before = len(seen_urls)
        items = parse_listing_page(html, seen_urls)
        if items:
            all_items.extend(items)
        after = len(seen_urls)
        logger.info("Collected so far: %d", after)

        if target_total and after >= target_total:
            logger.info("Reached target total %d. Stop pagination.", target_total)
            break

        # Find a next button.
        next_button = None
        for sel in ["a.ts-load-next:not(.ts-btn-disabled)", "text=下一頁"]:
            try:
                locator = page.locator(sel)
                if await locator.count() > 0:
                    next_button = locator.last
                    break
            except Exception:
                continue
        if next_button is None:
            logger.info("No usable next button found; stop pagination.")
            break

        # Click + poll until we truly see new items.
        progress = False
        for click_attempt in range(1, 4):
            try:
                await next_button.scroll_into_view_if_needed(timeout=1500)
            except Exception:
                pass
            try:
                await next_button.click(force=True, timeout=4000)
            except Exception:
                handle = await next_button.element_handle()
                if handle:
                    await page.evaluate("(el) => el.click()", handle)
            for _ in range(6):
                await page.wait_for_timeout(800)
                try:
                    await page.wait_for_load_state("networkidle", timeout=3000)
                except Exception:
                    pass
                probe_html = await page.content()
                probe_seen = set(seen_urls)
                parse_listing_page(probe_html, probe_seen)
                if len(probe_seen) > after:
                    progress = True
                    break
            if progress:
                stagnation = 0
                break
            logger.info("Pagination did not advance after click (%d/3)", click_attempt)

        if not progress:
            stagnation += 1
            if stagnation >= 2:
                logger.info("Stop pagination because there is no further progress.")
                break
        else:
            await page.wait_for_timeout(1200)

    return all_items


# -------------------------
# Delta generation
# -------------------------
def row_signature(row: pd.Series, cols: Sequence[str]) -> str:
    payload = "||".join(str(row.get(c, "") or "") for c in cols)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def build_delta(current_df: pd.DataFrame, previous_df: Optional[pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, int]]:
    key = "detail_url"
    compare_cols = [
        "title", "plan_source", "eligible_targets", "applicable_region", "grant_amount",
        "deadline_date", "deadline_text", "organizer_site_url", "official_organizer_site_url",
        "official_url_status",
    ]

    if previous_df is None or previous_df.empty or key not in previous_df.columns:
        new_df = current_df.copy()
        updated_df = current_df.iloc[0:0].copy()
        removed_df = current_df.iloc[0:0].copy()
        stats = {
            "current_count": int(len(current_df)),
            "previous_count": 0,
            "new_count": int(len(new_df)),
            "updated_count": 0,
            "removed_count": 0,
        }
        return new_df, updated_df, removed_df, stats

    curr = current_df.copy()
    prev = previous_df.copy()
    curr["__sig__"] = curr.apply(lambda r: row_signature(r, compare_cols), axis=1)
    prev["__sig__"] = prev.apply(lambda r: row_signature(r, compare_cols), axis=1)

    curr_map = curr.set_index(key)["__sig__"].to_dict()
    prev_map = prev.set_index(key)["__sig__"].to_dict()

    new_keys = [k for k in curr_map.keys() if k not in prev_map]
    removed_keys = [k for k in prev_map.keys() if k not in curr_map]
    updated_keys = [k for k in curr_map.keys() if k in prev_map and curr_map[k] != prev_map[k]]

    new_df = curr[curr[key].isin(new_keys)].drop(columns=["__sig__"], errors="ignore")
    updated_df = curr[curr[key].isin(updated_keys)].drop(columns=["__sig__"], errors="ignore")
    removed_df = prev[prev[key].isin(removed_keys)].drop(columns=["__sig__"], errors="ignore")
    stats = {
        "current_count": int(len(curr)),
        "previous_count": int(len(prev)),
        "new_count": int(len(new_df)),
        "updated_count": int(len(updated_df)),
        "removed_count": int(len(removed_df)),
    }
    return new_df, updated_df, removed_df, stats


# -------------------------
# Excel formatting
# -------------------------
def auto_fit_worksheet(ws, wrap_cols: Sequence[str] = ()):
    headers = [c.value for c in ws[1]]
    header_index = {h: i + 1 for i, h in enumerate(headers)}
    for i, header in enumerate(headers, start=1):
        max_len = len(str(header or ""))
        for row in ws.iter_rows(min_row=2, max_col=i, min_col=i):
            val = row[0].value
            if val is None:
                continue
            l = min(len(str(val)), 80)
            if l > max_len:
                max_len = l
        width = min(max(max_len + 2, 12), 45)
        ws.column_dimensions[get_column_letter(i)].width = width
        alignment = Alignment(vertical="top", wrap_text=header in wrap_cols)
        for cell in ws[get_column_letter(i)]:
            cell.alignment = alignment
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _rowify(item: Any) -> Dict[str, Any]:
    if isinstance(item, dict):
        return dict(item)
    if hasattr(item, "model_dump"):
        return item.model_dump()
    if hasattr(item, "dict"):
        return item.dict()
    try:
        return asdict(item)
    except Exception:
        raise TypeError(f"Unsupported row type for workbook export: {type(item)!r}")


def write_workbooks(detail_items: List[DetailItem], previous_df: Optional[pd.DataFrame]) -> Dict[str, int]:
    rows = [_rowify(x) for x in detail_items]
    detail_df = pd.DataFrame(rows)

    summary_cols = [
        "title", "detail_url", "plan_source", "eligible_targets", "applicable_region",
        "grant_amount", "deadline_date", "deadline_text",
        "topic_1", "topic_2", "topic_3", "topic_4", "topic_5",
        "organizer_site_url", "organizer_site_domain",
        "official_organizer_site_url", "official_organizer_domain",
        "official_url_status", "official_url_confidence",
    ]
    for c in summary_cols:
        if c not in detail_df.columns:
            detail_df[c] = ""
    summary_df = detail_df[summary_cols].copy()

    new_df, updated_df, removed_df, stats = build_delta(summary_df, previous_df)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="grants_summary", index=False)
        detail_df.to_excel(writer, sheet_name="grants_detail", index=False)
        pd.DataFrame([stats]).to_excel(writer, sheet_name="weekly_changes_summary", index=False)
        new_df.to_excel(writer, sheet_name="weekly_new_plans", index=False)
        updated_df.to_excel(writer, sheet_name="weekly_updated_plans", index=False)
        removed_df.to_excel(writer, sheet_name="weekly_removed_plans", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    auto_fit_worksheet(wb["grants_summary"], wrap_cols=[])
    auto_fit_worksheet(wb["grants_detail"], wrap_cols=["plan_background", "application_tips", "raw_text", "organizer_site_note"])
    auto_fit_worksheet(wb["weekly_new_plans"], wrap_cols=[])
    auto_fit_worksheet(wb["weekly_updated_plans"], wrap_cols=[])
    auto_fit_worksheet(wb["weekly_removed_plans"], wrap_cols=[])
    wb.save(OUTPUT_XLSX)

    with pd.ExcelWriter(DELTA_XLSX, engine="openpyxl") as writer:
        pd.DataFrame([stats]).to_excel(writer, sheet_name="changes_summary", index=False)
        new_df.to_excel(writer, sheet_name="new_plans", index=False)
        updated_df.to_excel(writer, sheet_name="updated_plans", index=False)
        removed_df.to_excel(writer, sheet_name="removed_plans", index=False)

    wb2 = load_workbook(DELTA_XLSX)
    auto_fit_worksheet(wb2["changes_summary"], wrap_cols=[])
    auto_fit_worksheet(wb2["new_plans"], wrap_cols=[])
    auto_fit_worksheet(wb2["updated_plans"], wrap_cols=[])
    auto_fit_worksheet(wb2["removed_plans"], wrap_cols=[])
    wb2.save(DELTA_XLSX)
    return stats


def load_previous_summary() -> Optional[pd.DataFrame]:
    if not Path(PREVIOUS_XLSX).exists():
        return None
    try:
        return pd.read_excel(PREVIOUS_XLSX, sheet_name="grants_summary")
    except Exception:
        try:
            return pd.read_excel(PREVIOUS_XLSX)
        except Exception:
            return None


# -------------------------
# Main
# -------------------------
async def main() -> None:
    cache = load_cache()
    previous_df = load_previous_summary()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(user_agent=UA, locale="zh-TW")
        list_page = await context.new_page()
        listing_items = await extract_all_listings(list_page)
        await list_page.close()
        logger.info("Found %d grants", len(listing_items))

        detail_page = await context.new_page()
        detail_items: List[DetailItem] = []
        for item in listing_items:
            detail = await extract_detail(detail_page, item)
            detail = resolve_by_title(detail, cache)
            # If detail lacked some listing values, backfill from listing.
            if not detail.plan_source:
                detail.plan_source = item.plan_source
            if not detail.eligible_targets:
                detail.eligible_targets = item.eligible_targets
            if not detail.grant_amount:
                detail.grant_amount = item.grant_amount
            if not detail.deadline_date:
                detail.deadline_date = item.deadline_date
            if not detail.topic_1:
                detail.topic_1 = item.topic_1
                detail.topic_2 = item.topic_2
                detail.topic_3 = item.topic_3
                detail.topic_4 = item.topic_4
                detail.topic_5 = item.topic_5
            detail_items.append(detail)
            await asyncio.sleep(0.3)
        await detail_page.close()
        await context.close()
        await browser.close()

    stats = write_workbooks(detail_items, previous_df)
    save_cache(cache)
    logger.info("Exported %d records to %s", len(detail_items), OUTPUT_XLSX)
    print(json.dumps(stats, ensure_ascii=False))


if __name__ == "__main__":
    asyncio.run(main())
